from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import gspread
from datetime import date
import os
import time
import json
import requests
from io import BytesIO
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager

# ---------------- SHARDING (env-driven) ---------------- #
SHARD_INDEX = int(os.getenv("SHARD_INDEX", "0"))
SHARD_STEP = int(os.getenv("SHARD_STEP", "1"))
START_INDEX = int(os.getenv("START_INDEX", "1"))
END_INDEX = int(os.getenv("END_INDEX", "2500"))
checkpoint_file = os.getenv("CHECKPOINT_FILE", "checkpoint_new_1.txt")
last_i = int(open(checkpoint_file).read()) if os.path.exists(checkpoint_file) else START_INDEX

# ---------------- SETUP ---------------- #
chrome_options = Options()
chrome_options.add_argument("--headless=new")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--remote-debugging-port=9222")

# ---------------- GOOGLE SHEETS AUTH ---------------- #
try:
    gc = gspread.service_account("credentials.json")
except Exception as e:
    print(f"Error loading credentials.json: {e}")
    exit(1)

sheet_data = gc.open('Tradingview Data Reel Experimental May').worksheet('Sheet5')

# ---------------- READ STOCK LIST FROM GITHUB EXCEL (BATCH MODE) ---------------- #
print("üì• Fetching stock list from GitHub Excel in batches...")

try:
    EXCEL_URL = "https://raw.githubusercontent.com/Lavit-sharma/stock_raja/main/Stock%20List.xlsx"
    response = requests.get(EXCEL_URL)
    response.raise_for_status()

    wb = load_workbook(BytesIO(response.content), read_only=True)
    ws = wb.active

    name_list = []
    company_list = []

    BATCH_READ_SIZE = 500  # read 500 rows at a time

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue  # skip header

        name = row[0] if len(row) > 0 and row[0] else ""
        company = row[4] if len(row) > 4 and row[4] else ""

        name_list.append(str(name))
        company_list.append(str(company))

        if row_idx % BATCH_READ_SIZE == 0:
            print(f"üìÑ Loaded {row_idx} rows so far...")

    print(f"‚úÖ Loaded {len(company_list)} companies (batch read).")

except Exception as e:
    print(f"‚ùå Error reading Excel from GitHub: {e}")
    exit(1)

current_date = date.today().strftime("%m/%d/%Y")

# ---------------- SCRAPER FUNCTION ---------------- #
def scrape_tradingview(company_url):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.set_window_size(1920, 1080)
    try:
        # LOGIN USING SAVED COOKIES
        if os.path.exists("cookies.json"):
            driver.get("https://www.tradingview.com/")
            with open("cookies.json", "r", encoding="utf-8") as f:
                cookies = json.load(f)
            for cookie in cookies:
                try:
                    cookie_to_add = {k: cookie[k] for k in ('name', 'value', 'domain', 'path') if k in cookie}
                    cookie_to_add['secure'] = cookie.get('secure', False)
                    cookie_to_add['httpOnly'] = cookie.get('httpOnly', False)
                    driver.add_cookie(cookie_to_add)
                except Exception:
                    pass
            driver.refresh()
            time.sleep(2)
        else:
            print("‚ö†Ô∏è cookies.json not found. Proceeding without login may limit data.")

        driver.get(company_url)
        WebDriverWait(driver, 45).until(
            EC.visibility_of_element_located((By.XPATH,
                '/html/body/div[2]/div/div[5]/div/div[1]/div/div[2]/div[1]/div[2]/div/div[1]/div[2]/div[2]/div[2]/div[2]/div'))
        )

        soup = BeautifulSoup(driver.page_source, "html.parser")
        values = [
            el.get_text().replace('‚àí', '-').replace('‚àÖ', '').strip()
            for el in soup.find_all("div", class_="valueValue-l31H9iuA apply-common-tooltip")
        ]
        return values

    except NoSuchElementException:
        print(f"Data element not found for URL: {company_url}")
        return []
    except Exception as e:
        print(f"An error occurred during scraping for {company_url}: {e}")
        return []
    finally:
        driver.quit()

# ---------------- MAIN LOOP WITH BATCH WRITING ---------------- #
buffer_rows = []
BATCH_WRITE_SIZE = 50  # write after every 50 entries

for i, company_url in enumerate(company_list[last_i:], last_i):
    if i < START_INDEX or i > END_INDEX:
        continue
    if i % SHARD_STEP != SHARD_INDEX:
        continue

    name = name_list[i] if i < len(name_list) else f"Row {i}"
    print(f"Scraping {i}: {name} | {company_url}")

    values = scrape_tradingview(company_url)
    if values:
        row = [name, current_date] + values
        buffer_rows.append(row)
        print(f"‚úÖ Added {name} to buffer ({len(buffer_rows)}/{BATCH_WRITE_SIZE})")

        if len(buffer_rows) >= BATCH_WRITE_SIZE:
            try:
                sheet_data.append_rows(buffer_rows, value_input_option='RAW')
                print(f"üìù Appended {len(buffer_rows)} rows to Google Sheet.")
                buffer_rows.clear()
                time.sleep(2)
            except Exception as e:
                print(f"‚ö†Ô∏è Failed to batch append: {e}")
                time.sleep(5)
    else:
        print(f"‚ö†Ô∏è Skipping {name}: No data scraped.")

    with open(checkpoint_file, "w") as f:
        f.write(str(i))

    time.sleep(1)

# Write any leftover buffered rows
if buffer_rows:
    try:
        sheet_data.append_rows(buffer_rows, value_input_option='RAW')
        print(f"üìù Final append of {len(buffer_rows)} rows.")
    except Exception as e:
        print(f"‚ö†Ô∏è Final append failed: {e}")
